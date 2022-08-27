from openpyxl import load_workbook
from openpyxl.styles import (
                        PatternFill, Border, Side,
                        Alignment, Font
                        )
from openpyxl.workbook import Workbook
from interservice_price_parser import parse_price
from new_price_converter import classify_literature


def get_data(file_name=None, price=None, sale=17):
    new_price = parse_price(file_name, price, sale)

    converted_price = classify_literature(new_price)

    return converted_price


def create_table_head(filename="test_us_price_1.xlsx"):
    wb = Workbook()
    page = wb.active

    page.merge_cells("A1:C1")
    page["A1"].value = 'ООО "УралСпециализация" поставляет литературу ведущих издательств России'
    page['A1'].alignment = Alignment(horizontal='center')
    page["A1"].font = Font(size=10)

    page["B2"].value = 'Адрес: 454048, г.Челябинск, ул.Доватора, 32 а (ул. Каменный лог, 11)'
    page['B2'].alignment = Alignment(horizontal='center')
    page["B2"].font = Font(size=8)

    page["B3"].value = 'Тел. (351)230-93-90, 8-951-111-05-73, Viber: +7-952-515-16-17'
    page['B3'].alignment = Alignment(horizontal='center')
    page["B3"].font = Font(size=8, color="000000FF")

    page["B4"].value = 'E-mail: sukniga@mail.ru, su-kniga74@mail.ru'
    page['B4'].alignment = Alignment(horizontal='center')
    page["B4"].font = Font(size=8, color="000000FF")

    page["B5"].value = 'Сайт: www.urfokniga.ru'
    page['B5'].alignment = Alignment(horizontal='center')
    page["B5"].font = Font(size=8, color="000000FF")

    page["B6"].value = 'Группа в Вконтакте:   https://vk.com/knigi_bez_problem'
    page['B6'].alignment = Alignment(horizontal='center')
    page["B6"].font = Font(size=8, color="000000FF")

    page.merge_cells("B7:B8")
    page["B7"].value = 'Предоставляются скидки к цене прайс-листа для постоянных покупателей или при разовой покупке.\n' \
                       ' ГАРАНТИЯ НИЗКОЙ ЦЕНЫ. Условия доставки оговариваются отдельно.  Обращайтесь! Будем рады сотрудничеству!'
    page['B7'].alignment = Alignment(horizontal='center')
    page["B7"].font = Font(size=10, color="0000FF00")

    page.merge_cells("D1:H1")
    page["D1"].value = "Для перехода к нужному классу, кликните на название"
    page["D1"].font = Font(size=8, color="00FF0000")

    page.column_dimensions['A'].width = 5
    page.column_dimensions['B'].width = 90
    page.column_dimensions['C'].width = 12

    links_cells = ["D2", "D3", "D4", "D5", "D6", "D7", "D8", "E2", "E3", "E4", "E5", "E6", "E7", "E8"]
    links = ["1 класс", "2 класс", "3 класс", "4 класс", "5 класс", "6 класс", "7 класс", "8 класс", "9 класс", "огэ", "10 класс", "11 класс", "егэ", "коррекция"]
    for index, cell in enumerate(links_cells):
        page[cell].value = links[index]
        page[cell].font = Font(size=8, color="000000FF")
        page[cell].alignment = Alignment(horizontal='center')

    page.freeze_panes = "E10"

    wb.save(filename)


def create_new_price(filename="test_us_price_1.xlsx", converted_price=None, sale=17):
    wb = load_workbook(filename=filename)
    page = wb.active

    headers = ["Класс", "Наименование", "ISBN", "Издательство", "Страниц", "Цена прайса", f"Цена прайса со скидкой {sale}%", "Заказ"]
    page.append(headers)

    for cls in converted_price.keys():
        if cls.isdigit():
            page.append([f"{cls}", f"{cls} Класс"])
        else:
            page.append([f"{cls}", f"{cls.upper()}"])
        for subject in converted_price[cls].keys():
            page.append([f"{cls}", f"{subject}"])
            rows = converted_price[cls][subject]
            for row in rows:
                page.append([row.get(key) for key in headers])

    wb.save(filename)


def arrange_new_price(filename="test_us_price_1.xlsx"):
    wb = load_workbook(filename=filename)
    page = wb.active

    border = Border(left=Side(border_style="medium"), right=Side(border_style="medium"), top=Side(border_style="medium"),
                    bottom=Side(border_style="medium"), diagonal=Side(border_style="medium"))

    table_len = len(page["A"])
    table_width = 8
    headers = ["0", "A", "B", "C", "D", "E", "F", "G", "H", "I"]

    def arrange_row(index, color, align=False, start_ind=1, bold=False):
        for ind in range(1, table_width + 1):
            page[f"{headers[ind]}{index}"].border = border

        for ind in range(start_ind, table_width + 1):
            page[f"{headers[ind]}{index}"].fill = PatternFill('solid', fgColor=color)
            page[f"{headers[ind]}{index}"].border = border
            if bold:
                page[f"{headers[ind]}{index}"].font = Font(size=10, bold=True)
            if align:
                page[f"{headers[ind]}{index}"].alignment = Alignment(horizontal='center')

    arrange_row(9, "CEA7E5", align=True)

    for index in range(10, table_len+1):
        if not page[f"C{index}"].value and not page[f"D{index}"].value and not page[f"E{index}"].value:
            if "Класс" in page[f"B{index}"].value or page[f"B{index}"].value in ["ОГЭ", "ЕГЭ", "КОР"]:
                arrange_row(index, "00FFFF00", bold=True)
            else:
                arrange_row(index, "FF9C06", start_ind=2, bold=True)
        else:
            for ind in range(1, table_width + 1):
                page[f"{headers[ind]}{index}"].border = border

    wb.save(filename)


